from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from pathlib import Path
import chardet

def fichier_du_chemin(chemin: str) -> str:
    """
    Extrait le nom du fichier sans l'extension à partir d'un chemin donné."""
    return Path(chemin).stem

def exist_file(file_path: str) -> bool:
    """
    Vérifie si un fichier existe à l'emplacement spécifié."""
    return Path(file_path).is_file()

def format_datetime(dt):
    """Formate un objet datetime en chaîne de caractères au format 'JJ-MM-AA_HH-MM'."""
    return dt.strftime("%d-%m-%y_%H-%M")

def nb_colonne(table):
    """Renvoie le nombre total de colonnes dans une table HTML, en tenant compte des colspan."""
    
    # Trouver le premier <tr> dans la table
    first_row = table.find('tr')  # Prend simplement le premier <tr> trouvé

    # Si aucune ligne n'est trouvée, retourner 0
    if first_row is None:
        return 0

    columns = first_row.find_all(['th', 'td'])  # Récupérer tous les éléments 'th' et 'td'

    # Compter le nombre total de colonnes en tenant compte du colspan
    total_columns = 0

    for col in columns:
        colspan = col.get('colspan', 1)  # Obtenir le colspan, par défaut, on prend 1
        total_columns += int(colspan)  # Ajouter à la somme totale des colonnes

    return total_columns

def excel_date_to_datetime(date):
    """
    Convertit une date Excel (sous forme de chaîne) en objet datetime."""
    excel_date = float(date.replace(",", "."))
    excel_epoch_start = datetime(1899, 12, 30)
    delta = timedelta(days=excel_date)
    return format_datetime(excel_epoch_start + delta)

def detect_encoding(file_path: str) -> str:
    """
    Détecte l'encodage d'un fichier en lisant les premiers 10 000 octets."""
    with open(file_path, 'rb') as file:
        rawdata = file.read(10000)
        result = chardet.detect(rawdata)
        return result['encoding']

def opti_html_to_xlsx(file_path: str, name=""):
    """
    Convertit un fichier HTML contenant une table en fichier Excel (.xlsx)."""
    if name == "":
        name = fichier_du_chemin(file_path)

    encoding_file = detect_encoding(file_path)

    with open(file_path, 'r', encoding=encoding_file) as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.find('table')
    nb_colonnes = nb_colonne(table)
    print(nb_colonnes)
    # Données à écrire dans Excel avec fusion à appliquer
    structured_data = []
    merge_instructions = []

    # Lire toutes les lignes
    rows = table.find_all('tr')
    for r_idx, row in enumerate(rows):
        structured_row = []
        col_idx = 0  # index réel de la colonne en tenant compte des fusions précédentes

        for cell in row.find_all(['th', 'td']):
            # Gestion des décalages causés par les cellules fusionnées
            while len(structured_row) <= col_idx:
                structured_row.append(None)

            # Lire la valeur de la cellule
            sdval = cell.get('sdval')
            sdnum = cell.get('sdnum')
            if sdval and sdnum == '1036;0;JJ/MM/AA HH:MM':
                value = excel_date_to_datetime(sdval)
            elif sdval:
                try:
                    value = float(sdval)
                except ValueError:
                    value = cell.text.strip()
            else:
                value = cell.text.strip()

            # Lire colspan
            colspan = int(cell.get("colspan", 1))
            rowspan = int(cell.get("rowspan", 1))

            structured_row[col_idx] = (value, colspan, rowspan)

            # Ajouter instruction de fusion si nécessaire
            if colspan > 1 or rowspan > 1:
                merge_instructions.append((r_idx + 1, col_idx + 1, rowspan, colspan))

            col_idx += colspan

        structured_data.append(structured_row)

    # Préparer le workbook
    result_path = f"{name}.xlsx"
    if exist_file(result_path):
        wb = load_workbook(result_path, data_only=True)
        if "opti" in wb.sheetnames:
            del wb["opti"]
        sheet = wb.create_sheet("opti")
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "opti"

    # Écriture des données 
    # Écriture des données 
    for r_idx, row in enumerate(structured_data, 1):
        c_idx = 1
        for col in range(nb_colonnes):
            if len(row)<=col:
                cell = None
                sheet.cell(row=r_idx, column=c_idx, value=sheet.cell(row=r_idx, column=c_idx-1).value)
            else:
                cell = row[col]
                if c_idx == 1 and cell[0] == None:
                    sheet.cell(row=r_idx, column=c_idx, value=None)
                elif cell is not None :
                    value, colspan, rowspan = cell
                    sheet.cell(row=r_idx, column=c_idx, value=value)
                else:
                    sheet.cell(row=r_idx, column=c_idx, value=sheet.cell(row=r_idx, column=c_idx-1).value)
                

            
            c_idx += 1

    print(c_idx)
    wb.save(result_path)


def convertir(chemin_source: str, chemin_destination: str):
    """
    Fonction publique pour lancer la conversion HTML → XLSX.
    """
    name = fichier_du_chemin(chemin_destination)
    opti_html_to_xlsx(chemin_source, name)

    

# opti_html_to_xlsx("data/Mesures.xls",taille_entete=2)
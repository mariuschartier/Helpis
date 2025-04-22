import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Optional









def coloriage_erreur_openpyxl(feuille,lig:int,col:int)->None:
    rouge = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") 
    feuille.cell(row=lig, column=col).fill = rouge


def coloriage_reset_openpyxl(feuille)->None:
    yellow_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row in feuille.iter_rows(min_row=feuille.max_row, max_col=feuille.max_column):  # ignorer les en-têtes (ligne 1)
        for cell in row:
            cell.fill = yellow_fill

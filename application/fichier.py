from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Optional

class Fichier:
    def __init__(self, chemin):
        self.chemin = chemin
        self.nom =Path(chemin).stem
        
    def __str__(self):
        return f"{self.chemin}(ligne{self.taille_entete})"
    
    
    
    def get_feuille(self, nom_col : str):
        df = pd.read_excel(self.chemin, sheet_name=nom_col, header=None, engine="openpyxl")
        return df
    
    def color_cell(self, lignes: list, col_index: int, feuille: str = None, couleur="FFC7CE"):
        """
        Colorie les cellules d'un fichier Excel aux lignes/colonnes spécifiées.
        
        :param fichier_excel: chemin vers le fichier Excel
        :param lignes: liste des indices de lignes à colorier (0-based par rapport au DataFrame, donc +1 pour Excel)
        :param col_index: index de colonne (0-based)
        :param feuille: nom de la feuille (facultatif)
        :param couleur: code couleur hex RGB (par défaut : rouge clair)
        """
        wb = load_workbook(self.chemin)
        ws = wb.active if feuille is None else wb[feuille]

        fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")

        for i in lignes:
            excel_row = i + 1  # Conversion 0-based → Excel (1-based)
            excel_col = col_index + 1
            ws.cell(row=excel_row, column=excel_col).fill = fill

        wb.save(self.chemin)
        print(f"✅ Cellules colorées dans {self.chemin}")





    def clear_all_cell_colors(self, feuille: Optional[str] = None):
        """
        Supprime toutes les couleurs de fond des cellules dans une feuille spécifique
        ou dans toutes les feuilles si aucune n’est précisée.

        :param fichier_excel: chemin vers le fichier Excel à modifier
        :param feuille: nom de la feuille à nettoyer (si None, toutes les feuilles sont traitées)
        """
        wb = load_workbook(self.chemin)

        feuilles_cibles = [wb[feuille]] if feuille else wb.worksheets

        for ws in feuilles_cibles:
            max_row = ws.max_row
            max_col = ws.max_column

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)

            print(f"✅ Couleurs effacées dans la feuille : '{ws.title}'")

        wb.save(self.chemin)
        print(f"📁 Fichier sauvegardé : {self.chemin}")
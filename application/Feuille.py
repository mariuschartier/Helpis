
from pathlib import Path


from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Optional

class Feuille:
    def __init__(self, fichier, nom, taille_entete:int):
        self.fichier = fichier
        self.nom = nom
        self.taille_entete = taille_entete
        self.df = self.fichier.get_feuille(self.nom)
        self.nb_ligne,  self.nb_colonne = self.df.shape
        
        
        #  les erreurs ont 3 valeurs de bases :
            # erreurs rouge pale dans les valeurs  et code 1
            # erreurs jaune coh√©rence entre les colonnes et code 3
            # il peux avoir plusieurs fois le meme code et des melange de code ce qui vas affecter la couleur
                # orange : + de 1 jaunes
                # rouge vif : + de 1 rouges
                # noire : orange + rouge vif
        self.erreurs = [[0 for _ in range(self.nb_colonne)] for _ in range(self.nb_ligne)]
        
    def __str__(self):
        return f"{self.chemin}(ligne{self.taille_entete})"
    
    
    
    def get_feuille(self):
        df = self.fichier.get_feuille(self.nom)
        return df
    
    
    def ajouts_erreur(self, lignes: list, col_index: int, code_erreur=1):
        for i in lignes:
            excel_row = i - self.taille_entete
            if 0 <= excel_row < self.nb_ligne:
                self.erreurs[excel_row + self.taille_entete][col_index] = code_erreur

    
    def color_cell(self, lignes: list, col_index: int, couleur="FFC7CE"):
        """
        Colorie les cellules d'un fichier Excel aux lignes/colonnes sp√©cifi√©es.
        
        :param fichier_excel: chemin vers le fichier Excel
        :param lignes: liste des indices de lignes √† colorier (0-based par rapport au DataFrame, donc +1 pour Excel)
        :param col_index: index de colonne (0-based)
        :param feuille: nom de la feuille (facultatif)
        :param couleur: code couleur hex RGB (par d√©faut : rouge clair)
        """
        wb = load_workbook(self.fichier.chemin)
        ws = wb[self.nom]

        fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")

        for i in lignes:
            excel_row = i + 1  # Conversion 0-based ‚Üí Excel (1-based)
            excel_col = col_index + 1
            ws.cell(row=excel_row, column=excel_col).fill = fill

        wb.save(self.fichier.chemin)
        print(f"‚úÖ Cellules color√©es dans {self.fichier.chemin}")


    


    def clear_all_cell_colors(self):
        """
        Supprime toutes les couleurs de fond des cellules dans une feuille sp√©cifique
        ou dans toutes les feuilles si aucune n‚Äôest pr√©cis√©e.

        :param fichier_excel: chemin vers le fichier Excel √† modifier
        :param feuille: nom de la feuille √† nettoyer (si None, toutes les feuilles sont trait√©es)
        """
        wb = load_workbook(self.fichier.chemin)

        feuilles_cibles = [wb[self.nom]] 

        for ws in feuilles_cibles:
            max_row = ws.max_row
            max_col = ws.max_column

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)

            print(f"‚úÖ Couleurs effac√©es dans la feuille : '{ws.title}'")

        wb.save(self.fichier.chemin)
        print(f"üìÅ Fichier sauvegard√© : {self.fichier.chemin}")
        
        
        
        
        
        
        
    def error_all_cell_colors(self):
        """
        Supprime toutes les couleurs de fond des cellules dans une feuille sp√©cifique
        ou dans toutes les feuilles si aucune n‚Äôest pr√©cis√©e.

        :param fichier_excel: chemin vers le fichier Excel √† modifier
        :param feuille: nom de la feuille √† nettoyer (si None, toutes les feuilles sont trait√©es)
        """
        wb = load_workbook(self.fichier.chemin)
        feuille_cible = wb[self.nom] 
        
        
        
        rouge="FFC7CE"
        jaune="F7DD24"
        rouge_ = PatternFill(start_color=rouge, end_color=rouge, fill_type="solid")
        jaune_ = PatternFill(start_color=jaune, end_color=jaune, fill_type="solid")

    

        for row in range( self.nb_ligne):
            for col in range( self.nb_colonne):

                if self.erreurs[row][col] == 1:
                    feuille_cible.cell(row=row+1, column=col+1).fill = rouge_
                elif self.erreurs[row][col] == 2:
                    feuille_cible.cell(row=row+1, column=col+1).fill = jaune_
        print(f"‚úÖ Erreurs couleur√©es dans la feuille : '{feuille_cible.title}'")

        wb.save(self.fichier.chemin)
        print(f"üìÅ Fichier sauvegard√© : {self.fichier.chemin}")
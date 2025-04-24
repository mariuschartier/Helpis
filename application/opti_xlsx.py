import openpyxl

def process_excel_data(input_file='results/Ponte Salle 1.xlsx', 
                       sheet_name='Données de ponte', 
                       output_file='results/processed_data.xlsx'):
    # Charger le fichier Excel avec les valeurs calculées
    wb = openpyxl.load_workbook(input_file, data_only=True)

    # Vérifier si la feuille existe
    if sheet_name not in wb.sheetnames:
        print(f"La feuille '{sheet_name}' n'existe pas dans le fichier.")
        return

    ws = wb[sheet_name]
    data = []

    # Lire toutes les lignes de données tout en conservant None pour les cellules vides
    for row in ws.iter_rows(min_row=1, values_only=True):
        # Ajouter chaque ligne comme une liste, tout en gérant les None
        data.append(list(row))

    # Traitement et gestion des None
    nb_colonnes = len(data[0]) if data else 0  # Nombre de colonnes de la première ligne
    structured_data = []

    for r_idx, row in enumerate(data, 1):
        structured_row = []
        for c_idx in range(nb_colonnes):
            if len(row) <= c_idx:
                cell = None
                value_to_set = structured_row[c_idx - 1] if c_idx > 0 else None
            else:
                cell = row[c_idx]
                if cell is None:
                    value_to_set = structured_row[c_idx - 1] if c_idx > 0 else None
                else:
                    value_to_set = cell

            structured_row.append(value_to_set)

        structured_data.append(structured_row)

    # Créer un nouveau classeur et une nouvelle feuille
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "Données traitées"

    # Écrire les données structurées dans le fichier
    for r_idx, row in enumerate(structured_data, 1):
        for c_idx, value in enumerate(row, 1):
            new_ws.cell(row=r_idx, column=c_idx, value=value)

    # Enregistrer le fichier Excel
    new_wb.save(output_file)
    print(f"Les données ont été enregistrées dans '{output_file}'.")

    # Afficher les 100 premières lignes de la structure finale
    for index, row in enumerate(structured_data[:100]):  # Limiter à 100 lignes
        print(f"Ligne {index + 1}: {row}")  # Afficher chaque ligne

# Exemple d'utilisation
# process_excel_data()
    
    
    
    
# # Exemple d'utilisation
# input_file = 'results/Ponte Salle 1.xlsx'
# output_file = 'output.xlsx'
# sheet_name = "Données de ponte" # Remplacez par le nom de votre feuille
# process_xlsx(input_file, output_file, sheet_name)
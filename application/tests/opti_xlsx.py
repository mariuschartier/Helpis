from structure.Fichier import Fichier
from structure.Feuille import Feuille
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import win32com.client as win32


import os
import win32com.client as win32
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def get_excel_value(file_path, sheet_name):
    """
    Récupère les valeurs d'une feuille Excel.

    :param file_path: Chemin du fichier Excel
    :param sheet_name: Nom de la feuille
    :return: Liste de listes représentant les valeurs de la feuille
    """
    print(f"Chemin fourni : {file_path}")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Le fichier '{file_path}' est introuvable.")

    data = []

    try:
        # Démarrer une instance d'Excel
        excel = win32.Dispatch("Excel.Application")

        # Ouvrir le classeur
        workbook = excel.Workbooks.Open(file_path, ReadOnly=True)

        if not workbook:
            raise ValueError("Le classeur n'a pas pu être ouvert.")

        # Vérifie si la feuille existe
        if not any(sheet.Name == sheet_name for sheet in workbook.Sheets):
            raise ValueError(f"La feuille '{sheet_name}' n'existe pas.")

        # Accéder à la feuille
        sheet = workbook.Sheets(sheet_name)
        for row in range(1, sheet.UsedRange.Rows.Count + 1):
            row_data = []
            for col in range(1, sheet.UsedRange.Columns.Count + 1):
                cell = sheet.Cells(row, col)
                value = cell.Value

                # Si la cellule contient un datetime
                if isinstance(value, datetime):
                    # Supprimer tzinfo si présent
                    if value.tzinfo is not None:
                        value = value.replace(tzinfo=None)
                    # Formater la date selon le format souhaité
                    value = value.strftime("%Y-%m-%d")  # <-- change ici si tu veux un autre format

                row_data.append(value)
            data.append(row_data)

        # Fermer le classeur sans enregistrer
        workbook.Close(SaveChanges=False)
        excel.Quit()

        return data

    except Exception as e:
        print(f"Erreur lors de la lecture du fichier Excel : {e}")
        if 'workbook' in locals() and workbook:
            workbook.Close(SaveChanges=False)
        if 'excel' in locals():
            excel.Quit()
        raise


def detect_col_row_span(file_path, sheet_name):
    """
    Détecte les cellules fusionnées dans une feuille Excel.
    Retourne une liste de tuples : (ligne_depart, colonne_depart, rowspan, colspan)
    sans doublons pour les plages fusionnées.
    """
    merged_cells = []
    seen_areas = set()

    try:
        excel = win32.Dispatch("Excel.Application")
        workbook = excel.Workbooks.Open(file_path, ReadOnly=True)
        sheet = workbook.Sheets(sheet_name)

        max_row = sheet.UsedRange.Rows.Count
        max_col = sheet.UsedRange.Columns.Count

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.Cells(row, col)
                if cell.MergeCells:
                    merge_area = cell.MergeArea
                    start_row = merge_area.Row
                    start_col = merge_area.Column
                    key = (start_row, start_col, merge_area.Rows.Count, merge_area.Columns.Count)
                    if key not in seen_areas:
                        seen_areas.add(key)
                        merged_cells.append((start_row, start_col, merge_area.Rows.Count, merge_area.Columns.Count))

        workbook.Close(SaveChanges=False)
        excel.Quit()
        return merged_cells

    except Exception as e:
        print(f"Erreur lors de la détection des cellules fusionnées : {e}")
        if 'workbook' in locals() and workbook:
            workbook.Close(SaveChanges=False)
        if 'excel' in locals():
            excel.Quit()
        raise

def process_and_format_excel(input_file, sheet_name, output_file):
    """
    Lit, détecte les cellules fusionnées et formate les données d'un fichier Excel.
    Les valeurs des cellules fusionnées sont appliquées uniquement sur les lignes correspondant au rowspan,
    et les autres colonnes de la plage fusionnée sont remplies avec None.

    :param input_file: Chemin du fichier Excel d'entrée
    :param sheet_name: Nom de la feuille à traiter
    :param output_file: Chemin du fichier Excel de sortie
    """
    # Lire les données de la feuille
    data = get_excel_value(input_file, sheet_name)

    # Détecter les cellules fusionnées
    merged_cells = detect_col_row_span(input_file, sheet_name)
    for row, col, rowspan, colspan in merged_cells:
        value = data[row - 1][col - 1]  # valeur dans la cellule fusionnée
        # print(f"Cellule fusionnée détectée à ({row}, {col}) avec rowspan={rowspan} et colspan={colspan}. Valeur : {value}")
        for c in range(col, col + colspan):  # pour chaque ligne du rowspan
            for r in range(row, row + rowspan):  # pour chaque colonne de la plage

                if r == row:  # dans la colonne de départ de la fusion
                    data[r - 1][c - 1] = value
                    
                else:  # autres colonnes de la plage
                    data[r - 1][c - 1] = None


    # Créer un nouveau classeur et une nouvelle feuille
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Données traitées"

    # Écrire les données dans la nouvelle feuille
    for r_idx, row in enumerate(data, 1):
        for c_idx, value in enumerate(row, 1):
            new_ws.cell(row=r_idx, column=c_idx, value=value)

    # Enregistrer le fichier formaté
    new_wb.save(output_file)
    print(f"Les données ont été enregistrées dans '{output_file}'.")

def determine_jour(date_str):
    # Parser la date
    dt = datetime.strptime(date_str, "%d/%m/%y %H:%M")
    
    # Extraire jour du mois, mois et année en format 4 chiffres
    jour_mois = dt.strftime("%d")  # exemple : '15'
    mois = dt.strftime("%m")       # exemple : '08'
    annee = dt.strftime("%Y")      # exemple : '2023'
    
    return jour_mois, mois, annee

def moyenne_par_jour(feuille,output_file ,date_col=0):
    wb = Workbook()
    ws = wb.active
    ws.title = "moyenne"

    nb_lignes = feuille.nb_ligne
    nb_colonnes = min(feuille.nb_colonne, 21)

    # Copier les lignes d'entête
    for row_idx in range(feuille.entete.taille_entete):
        for col_idx in range(nb_colonnes):
            valeur = feuille.df.iloc[row_idx, col_idx]
            ws.cell(row=row_idx + 1, column=col_idx + 1, value=valeur)

    # Regrouper les lignes par jour
    jours_dict = {}
    for idx in range(feuille.debut_data, nb_lignes):
        date_cell = feuille.df.iloc[idx, date_col]
        if pd.isna(date_cell):
            continue
    
        try:
            jour, mois, annee = determine_jour(date_cell)
        except Exception:
            continue  # Ignorer si la cellule n'est pas une date valide

        clef_jour = f"{jour}/{mois}/{annee}"
        if clef_jour not in jours_dict:
            jours_dict[clef_jour] = {"rows": [], "date": clef_jour}
        jours_dict[clef_jour]["rows"].append(idx)

    # Calculer la moyenne pour chaque jour
    ligne_resultat = feuille.debut_data
    for clef_jour, info in sorted(jours_dict.items(), key=lambda x: datetime.strptime(x[1]["date"], "%d/%m/%Y")):
        
        rows = info["rows"]
        moyennes = []
        for col_index in range(nb_colonnes):
            somme = 0
            count = 0
            for row in rows:
                val = feuille.df.iloc[row, col_index]
                if isinstance(val, (int, float)):
                    somme += val
                    count += 1
            moyenne = somme / count if count > 0 else None
            moyennes.append(moyenne)

        # écrire le jour en 1re colonne
        ws.cell(row=ligne_resultat, column=1, value=clef_jour)
        # écrire les moyennes à partir de la 2ème colonne
        for idx, moyenne in enumerate(moyennes, start=1):
            ws.cell(row=ligne_resultat, column=idx, value=moyenne)

        ligne_resultat += 1

    # Enregistrer le fichier Excel
    wb.save(output_file)
    print(f"Les données ont été enregistrées dans '{output_file}'.")


def determine_semaine(date_str):
    # Parser la date
    dt = datetime.strptime(date_str, "%d/%m/%y %H:%M")

    # Extraire annee et semaine ISO
    annee, semaine, _ = dt.isocalendar()
    return f"{annee}-S{semaine:02d}", dt

def moyenne_par_semaine(feuille, output_file,date_col=0):
    wb = Workbook()
    ws = wb.active
    ws.title = "moyenne"

    nb_lignes = feuille.nb_ligne
    nb_colonnes = min(feuille.nb_colonne, 21)

    # Copier les lignes d'entête
    for row_idx in range(feuille.entete.taille_entete):
        for col_idx in range(nb_colonnes):
            valeur = feuille.df.iloc[row_idx, col_idx]
            ws.cell(row=row_idx + 1, column=col_idx + 1, value=valeur)

    # Regrouper les lignes par semaine
    semaines_dict = {}
    for idx in range(feuille.debut_data, nb_lignes):
        date_cell = feuille.df.iloc[idx, date_col]
        if pd.isna(date_cell):
            continue
        if isinstance(date_cell, datetime):
            dt = date_cell
        else:
            try:
                dt = datetime.strptime(str(date_cell), "%d/%m/%y %H:%M")
            except Exception:
                continue  # Ignorer si la cellule n'est pas une date valide

        clef_semaine, dt_obj = determine_semaine(dt.strftime("%d/%m/%y %H:%M"))

        if clef_semaine not in semaines_dict:
            semaines_dict[clef_semaine] = {"rows": [], "date": dt_obj}
        semaines_dict[clef_semaine]["rows"].append(idx)

    # Calculer la moyenne pour chaque semaine
    ligne_resultat = feuille.debut_data
    for clef_semaine, info in sorted(semaines_dict.items(), key=lambda x: x[1]["date"]):
        rows = info["rows"]
        moyennes = []
        for col_index in range(nb_colonnes):
            somme = 0
            count = 0
            for row in rows:
                val = feuille.df.iloc[row, col_index]
                if isinstance(val, (int, float)):
                    somme += val
                    count += 1
            moyenne = somme / count if count > 0 else None
            moyennes.append(moyenne)

        # écrire la semaine en 1re colonne
        ws.cell(row=ligne_resultat, column=1, value=clef_semaine)
        # écrire les moyennes à partir de la 2ème colonne
        for idx, moyenne in enumerate(moyennes, start=1):
            ws.cell(row=ligne_resultat, column=idx, value=moyenne)

        ligne_resultat += 1

    # Enregistrer le fichier Excel
    wb.save(output_file)
    print(f"Les données ont été enregistrées dans '{output_file}'.")




def entete_une_ligne(feuille: Feuille, output_file):
    """
    Génère un fichier Excel où l'entête (même si elle est sur plusieurs lignes) est ramenée sur une seule ligne.
    Chaque colonne d'entête est concaténée .
    """
    wb = feuille.one_line_header_openpyxl()
    wb.save(output_file)
    print(f"Les données ont été enregistrées dans '{output_file}'.")


    
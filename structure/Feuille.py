
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Optional
import pandas as pd
from structure.Entete import Entete
from structure.Fichier import Fichier
from openpyxl import Workbook

class Feuille:
    """ Représente une feuille d'un fichier Excel, avec ses données et son entête.
    Cette classe permet de manipuler les données d'une feuille, d'ajouter des erreurs, de colorier des cellules,
    et de gérer l'entête de la feuille."""
    def __init__(self, fichier : Fichier, nom, debut_data = None, fin_data = None):
        """ Initialise la feuille avec le fichier Excel, le nom de la feuille,
        et les indices de début et de fin des données."""
        self.fichier = fichier
        self.nom = nom
        self.df = self.fichier.get_feuille(self.nom)
        self.nb_ligne,  self.nb_colonne = self.df.shape
        self.entete =  Entete(self)

        if debut_data == None:
            self.debut_data = self.entete.taille_entete + 1
        else:
            self.debut_data = debut_data


        if fin_data == None:
            self.fin_data = len(self.df.index) - 1
        else:
            self.fin_data = fin_data


        
        #  les erreurs ont 3 valeurs de bases :
            # erreurs rouge pale dans les valeurs  et code 1
            # erreurs jaune cohérence entre les colonnes et code 3
            # il peux avoir plusieurs fois le meme code et des melange de code ce qui vas affecter la couleur
                # orange : + de 1 jaunes
                # rouge vif : + de 1 rouges
                # noire : orange + rouge vif
        self.erreurs = [[0 for _ in range(self.nb_colonne)] for _ in range(self.nb_ligne)]
        
    def __str__(self):
        return f"{self.chemin}(ligne{self.entete.taille_entete})"
    
    def get_feuille(self):
        """ Retourne le DataFrame de la feuille."""
        df = self.fichier.get_feuille(self.nom)
        return df
    
    
    def ajouts_erreur(self, lignes: list, col_index: int, code_erreur=1):
        """
        Ajoute une erreur à la feuille aux lignes et colonnes spécifiées."""
        for i in lignes:
            excel_row = i - self.entete.taille_entete
            if 0 <= excel_row < self.nb_ligne:
                self.erreurs[excel_row + self.entete.taille_entete][col_index] = code_erreur
    
    def color_cell(self, lignes: list, col_index: int, couleur="FFC7CE"):
        """
        Colorie les cellules d'un fichier Excel aux lignes/colonnes spécifiées.
        
        :param fichier_excel: chemin vers le fichier Excel
        :param lignes: liste des indices de lignes à colorier (0-based par rapport au DataFrame, donc +1 pour Excel)
        :param col_index: index de colonne (0-based)
        :param feuille: nom de la feuille (facultatif)
        :param couleur: code couleur hex RGB (par défaut : rouge clair)
        """
        wb = load_workbook(self.fichier.chemin)
        ws = wb[self.nom]

        fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")

        for i in lignes:
            excel_row = i + 1  # Conversion 0-based → Excel (1-based)
            excel_col = col_index + 1
            ws.cell(row=excel_row, column=excel_col).fill = fill

        wb.save(self.fichier.chemin)
        print(f"✅ Cellules colorées dans {self.fichier.chemin}")

    def clear_all_cell_colors(self):
        """
        Supprime toutes les couleurs de fond des cellules dans une feuille spécifique
        ou dans toutes les feuilles si aucune n’est précisée.

        :param fichier_excel: chemin vers le fichier Excel à modifier
        :param feuille: nom de la feuille à nettoyer (si None, toutes les feuilles sont traitées)
        """
        wb = load_workbook(self.fichier.chemin)

        feuilles_cibles = [wb[self.nom]] 

        for ws in feuilles_cibles:
            max_row = ws.max_row
            max_col = ws.max_column

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)

            print(f"✅ Couleurs effacées dans la feuille : '{ws.title}'")

        wb.save(self.fichier.chemin)
        print(f"📁 Fichier sauvegardé : {self.fichier.chemin}")
            
    def error_all_cell_colors(self):
        """
        Supprime toutes les couleurs de fond des cellules dans une feuille spécifique
        ou dans toutes les feuilles si aucune n’est précisée.

        :param fichier_excel: chemin vers le fichier Excel à modifier
        :param feuille: nom de la feuille à nettoyer (si None, toutes les feuilles sont traitées)
        """
        wb = load_workbook(self.fichier.chemin)
        feuille_cible = wb[self.nom] 
        
        
        
        rouge="FFC7CE"
        jaune="F7DD24"
        rouge_ = PatternFill(start_color=rouge, end_color=rouge, fill_type="solid")
        jaune_ = PatternFill(start_color=jaune, end_color=jaune, fill_type="solid")

    

        for row in range( self.nb_ligne):
            for col in range( self.nb_colonne):

                if self.erreurs[row][col] == 1:
                    feuille_cible.cell(row=row+1, column=col+1).fill = rouge_
                elif self.erreurs[row][col] == 2:
                    feuille_cible.cell(row=row+1, column=col+1).fill = jaune_
        print(f"✅ Erreurs couleurées dans la feuille : '{feuille_cible.title}'")

        wb.save(self.fichier.chemin)
        print(f"📁 Fichier sauvegardé : {self.fichier.chemin}")
                
    def one_line_header_pandas(self)->pd.DataFrame:
        """ Retourne un DataFrame avec l'entête fusionnée en une seule ligne,
        suivi des données de la feuille."""
        # Obtenir la liste représentant l'entête fusionnée
        entete_list = self.entete.une_ligne()

        # Convertir la liste en DataFrame avec une seule ligne
        entete_df = pd.DataFrame([entete_list])

        # Extraire les données
        data = self.df.iloc[self.debut_data:self.fin_data].reset_index(drop=True)

        # Concaténer l'entête et les données
        resultat = pd.concat([entete_df, data], ignore_index=True)

        return resultat
            
    def one_line_header_openpyxl(self, filename=None)->Workbook:

        """ Retourne un Workbook avec l'entête fusionnée en une seule ligne,
        suivi des données de la feuille."""
        # Obtenir l'entête sous forme de liste
        entete_list = self.entete.une_ligne()
        # Extraire les données
        data = self.df.iloc[self.debut_data+1:self.fin_data].reset_index(drop=True)

        # Créer un nouveau workbook ou ouvrir un existant
        wb = Workbook()
        ws = wb.active  # Utiliser la feuille active

        # Écrire la ligne d'entête fusionnée en première ligne
        for col_num, header_value in enumerate(entete_list, start=1):
            ws.cell(row=1, column=col_num, value=header_value)

        # Écrire les données à partir de la deuxième ligne
        for row_idx, row in data.iterrows():
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
        return wb
        # Enregistrer le fichier
        # wb.save(filename)


    def maj_feuille(self,fichier: Fichier, nom: str, debut_data: Optional[int] = None, fin_data: Optional[int] = None):
        """
        Met à jour ou crée une feuille dans un fichier Excel.
        
        :param fichier: Instance de la classe Fichier contenant le chemin du fichier Excel.
        :param nom: Nom de la feuille à mettre à jour ou créer.
        :param debut_data: Indice de début des données (optionnel).
        :param fin_data: Indice de fin des données (optionnel).
        :return: Instance de la classe Feuille mise à jour.
        """
        if isinstance(fichier, Fichier):
            self.fichier = fichier
        if  isinstance(nom, str) :
            self.nom = nom
        if isinstance(debut_data, int):
            self.debut_data = debut_data
        if  isinstance(fin_data, int):
            self.fin_data = fin_data        
        
        

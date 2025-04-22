# -*- coding: utf-8 -*-
"""
Created on Mon Apr  7 14:51:28 2025

@author: m.chartierlegoff
"""

from openpyxl import Workbook

def moyenne_block(sheet, taille_block=10):
    wb = Workbook()
    ws = wb.active
    ws.title = "moyenne"

    nb_lignes = sheet.max_row
    nb_colonnes = min(sheet.max_column, 21)  # On limite à 21 colonnes max
    nb_block = nb_lignes // taille_block
    reste = nb_lignes % taille_block

    def calcule_moyenne_bloc(start_row, end_row):
        moyennes = []
        for col_index in range(1, nb_colonnes + 1):
            somme = 0
            count = 0
            for row in range(start_row, end_row + 1):
                val = sheet.cell(row=row, column=col_index).value
                if isinstance(val, (int, float)):
                    somme += val
                    count += 1
            moyenne = somme / count if count > 0 else None
            moyennes.append(moyenne)
        return moyennes

    ligne_resultat = 1
    # Traitement des blocs complets
    for i in range(nb_block):
        start_row = i * taille_block + 1
        end_row = start_row + taille_block - 1
        moyennes = calcule_moyenne_bloc(start_row, end_row)
        for col_index, moyenne in enumerate(moyennes, start=1):
            ws.cell(row=ligne_resultat, column=col_index, value=moyenne)
        ligne_resultat += 1

    # Traitement du bloc restant
    if reste > 0:
        start_row = nb_block * taille_block + 1
        end_row = nb_lignes
        moyennes = calcule_moyenne_bloc(start_row, end_row)
        for col_index, moyenne in enumerate(moyennes, start=1):
            ws.cell(row=ligne_resultat, column=col_index, value=moyenne)

    wb.save("moyenne.xlsx")
    print("Fichier moyenne.xlsx sauvegardé.")
